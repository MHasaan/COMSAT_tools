"""
Timetable PDF Scraper for aSc Timetables (CUI Lahore)

Extracts class schedule data from aSc Timetables PDFs (CUI Lahore).
Each page contains one section's weekly grid with 24 time slots and 7 day rows.
Slot duration is auto-detected from the PDF header (e.g. 30-min or 20-min slots).

Phase 1: Extract all entries (Section, Day, Slots, Time, Subject, Room)
Phase 2: Pivot by room -> Excel workbook with one sheet per room (grid view)
         with conflict detection (red-highlighted cells)

Usage:
    python main.py                          # Process all pages, default PDF
    python main.py <pdf_path>               # Process all pages of given PDF
    python main.py <pdf_path> <max_pages>   # Process first N pages (0 = all)
    python main.py <pdf_path> <max_pages> <output.xlsx>
"""

import glob
import csv
import sys
import os
import re
from collections import defaultdict

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DAYS = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]


def _natural_sort_key(text):
    """Sort key that handles embedded numbers naturally (A-2 before A-10)."""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text)]


DEFAULT_INPUT_DIR = os.path.join(
    os.path.dirname(__file__), "..", "Data", "Input", "schedule"
)

DEFAULT_OUTPUT = os.path.join(
    os.path.dirname(__file__), "..", "Data", "Output", "schedule", "room_timetables.xlsx"
)

DEFAULT_CSV_OUTPUT = os.path.join(
    os.path.dirname(__file__), "..", "Data", "Output", "schedule", "timetable_processed.csv"
)


def _get_latest_pdf(directory):
    """Return the most recently modified PDF in *directory*, or None."""
    pdfs = glob.glob(os.path.join(directory, "*.pdf"))
    if not pdfs:
        return None
    return max(pdfs, key=os.path.getmtime)

# ---------------------------------------------------------------------------
# Styles for Excel
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DAY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
DAY_FONT = Font(bold=True, size=11)
CONFLICT_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
CONFLICT_FONT = Font(bold=True, color="FFFFFF", size=9)
CELL_FONT = Font(size=9)
CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="center", horizontal="center")
TITLE_FONT = Font(bold=True, size=14)

# ---------------------------------------------------------------------------
# Time-slot detection helpers
# ---------------------------------------------------------------------------


def _to_24h(time_str, prev_24h_hour=None):
    """
    Convert an ambiguous 12-hour time string (e.g. '1:00', '8:30') to 24-hour
    HH:MM format. Uses prev_24h_hour (the previous slot's 24-hour hour) to
    resolve AM/PM ambiguity:
      - Hours 12 stay as 12
      - Hours 1-7 become 13-19 (PM) when they follow an hour >= 12
      - Hours 8-11 are AM on first encounter, PM if they follow an hour >= 15
    """
    parts = time_str.strip().split(":")
    h = int(parts[0])
    m = int(parts[1]) if len(parts) > 1 else 0

    if prev_24h_hour is not None:
        # After noon: 1-7 are always PM (13-19)
        if 1 <= h <= 7 and prev_24h_hour >= 12:
            h += 12
        # 8-11 appearing after, say, 15:xx means PM (20-23)
        elif 8 <= h <= 11 and prev_24h_hour >= 15:
            h += 12
    # First slot: 8-11 assumed AM, 12 stays 12, 1-7 assumed PM
    else:
        if 1 <= h <= 7:
            h += 12

    return f"{h:02d}:{m:02d}"


def detect_slot_times(pdf):
    """
    Read the first few pages of *pdf* to discover the slot grid.
    Returns:
        slot_times:      dict[slot_num] -> (start_12h, end_12h)   (original strings)
        slot_times_24h:  dict[slot_num] -> "HH:MM"                (24-hour start time)
        num_slots:       int  (e.g. 24)
    """
    # Try up to the first 5 pages to find a valid header row
    for pi in range(min(5, len(pdf.pages))):
        table = pdf.pages[pi].extract_table()
        if not table or len(table) < 2:
            continue
        for row in table:
            # Look for the row whose cells match 'slot_num\nstart\nend'
            for cell in row[1:5]:
                if cell and re.match(r"^\d+\n", cell):
                    # Found header row — parse all slots
                    raw = {}  # slot_num -> (start_12h, end_12h)
                    for ci in range(1, len(row)):
                        c = row[ci]
                        if not c:
                            continue
                        parts = c.strip().split("\n")
                        if len(parts) >= 3:
                            raw[int(parts[0])] = (parts[1].strip(), parts[2].strip())

                    if not raw:
                        continue

                    num_slots = max(raw.keys())

                    # Build 24-hour mapping by walking slots in order
                    slot_times = {}      # slot -> (start_12h, end_12h)
                    slot_times_24h = {}  # slot -> "HH:MM"
                    prev_h = None
                    for sn in range(1, num_slots + 1):
                        if sn not in raw:
                            continue
                        s12, e12 = raw[sn]
                        s24 = _to_24h(s12, prev_h)
                        slot_times[sn] = (s12, e12)
                        slot_times_24h[sn] = s24
                        prev_h = int(s24.split(":")[0])

                    # Detect slot duration for logging
                    if len(slot_times_24h) >= 2:
                        t1 = list(slot_times_24h.values())[0]
                        t2 = list(slot_times_24h.values())[1]
                        h1, m1 = map(int, t1.split(":"))
                        h2, m2 = map(int, t2.split(":"))
                        duration = (h2 * 60 + m2) - (h1 * 60 + m1)
                    else:
                        duration = 0

                    print(f"Detected {num_slots} slots, {duration}-minute intervals")
                    print(f"  Range: {slot_times_24h[1]} – {slot_times_24h[num_slots]}")
                    return slot_times, slot_times_24h, num_slots

    # Fallback: should not happen with a valid timetable PDF
    raise RuntimeError("Could not detect time slots from PDF header")


# ---------------------------------------------------------------------------
# PDF Extraction Helpers
# ---------------------------------------------------------------------------


def parse_time_slots(header_row):
    """
    Parse the header row to build a mapping: column_index -> (slot_number, start_time, end_time).
    Header cells look like: '1\\n8:30\\n9:00'
    """
    slots = {}
    for col_idx in range(1, len(header_row)):
        cell = header_row[col_idx]
        if not cell:
            continue
        parts = cell.strip().split("\n")
        if len(parts) >= 3:
            slot_num = int(parts[0])
            start_time = parts[1].strip()
            end_time = parts[2].strip()
            slots[col_idx] = (slot_num, start_time, end_time)
    return slots


def parse_cell(cell_text):
    """
    Parse a cell's text into (subject, room).
    Last line = room, rest = subject. If only 1 line, subject only.
    """
    lines = [ln.strip() for ln in cell_text.strip().split("\n") if ln.strip()]
    if len(lines) == 0:
        return None, None
    if len(lines) == 1:
        return lines[0], ""
    room = lines[-1]
    subject = " ".join(lines[:-1])
    return subject, room


def extract_section_name(page_text):
    """
    Extract the section identifier from the page text.
    Patterns: 'FA25-CHE-A (Semester 2)' or '1 FA24-CHE-A (Semester 4)'.
    """
    match = re.search(
        r"(?:\d+\s+)?((?:FA|SP)\d{2}-\S+\s*\(Semester\s*\d+\))", page_text
    )
    if match:
        return match.group(1).strip()
    lines = [ln.strip() for ln in page_text.split("\n") if ln.strip()]
    if len(lines) >= 2:
        return lines[1]
    return "Unknown Section"


def extract_page_entries(page, page_index):
    """
    Extract all timetable entries from a single page.
    Returns list of dicts: {section, day, start_slot, end_slot, slots, time_range, subject, room}
    """
    text = page.extract_text() or ""
    section = extract_section_name(text)

    table = page.extract_table()
    if not table or len(table) < 3:
        return []

    # Find header row (contains slot numbers like '1\n8:30\n9:00')
    header_row_idx = None
    for ri, row in enumerate(table):
        for cell in row[1:5]:
            if cell and re.match(r"^\d+\n", cell):
                header_row_idx = ri
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return []

    slots = parse_time_slots(table[header_row_idx])
    if not slots:
        return []

    entries = []

    for ri in range(header_row_idx + 1, len(table)):
        row = table[ri]
        day = (row[0] or "").strip()
        if day not in DAYS:
            continue

        col = 1
        max_col = len(row)
        while col < max_col:
            cell = row[col]
            if cell is None or cell == "":
                col += 1
                continue

            # Determine span by counting trailing None cells
            start_col = col
            end_col = col
            k = col + 1
            while k < max_col and row[k] is None:
                end_col = k
                k += 1

            subject, room = parse_cell(cell)
            if subject is None:
                col = k
                continue

            # Map columns to slot numbers and time range
            if start_col in slots and end_col in slots:
                start_slot = slots[start_col][0]
                end_slot = slots[end_col][0]
                start_time = slots[start_col][1]
                end_time = slots[end_col][2]
            elif start_col in slots:
                start_slot = slots[start_col][0]
                end_slot = start_slot
                start_time = slots[start_col][1]
                end_time = slots[start_col][2]
            else:
                start_slot = start_col
                end_slot = end_col
                start_time = "?"
                end_time = "?"

            slot_str = str(start_slot) if start_slot == end_slot else f"{start_slot}-{end_slot}"
            time_range = f"{start_time} - {end_time}"

            entries.append({
                "section": section,
                "day": day,
                "start_slot": start_slot,
                "end_slot": end_slot,
                "slots": slot_str,
                "time_range": time_range,
                "subject": subject,
                "room": room,
            })

            col = k

    return entries


# ---------------------------------------------------------------------------
# Room Pivot & Conflict Detection
# ---------------------------------------------------------------------------


def build_room_grids(all_entries, slot_times, num_slots):
    """
    Pivot entries by room. For each room, build a grid:
        grid[day][slot_number] = list of (section, subject)

    Returns:
        room_grids: dict[room] -> dict[day] -> dict[slot] -> list of (section, subject)
        conflicts:  list of {room, day, slot, time, entries}
    """
    room_grids = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    conflicts = []

    for entry in all_entries:
        room = entry["room"]
        if not room:
            continue

        day = entry["day"]
        section = entry["section"]
        subject = entry["subject"]

        for slot in range(entry["start_slot"], entry["end_slot"] + 1):
            room_grids[room][day][slot].append((section, subject))

    # Detect conflicts: any slot with more than 1 entry
    for room in room_grids:
        for day in DAYS:
            for slot in range(1, num_slots + 1):
                occupants = room_grids[room][day][slot]
                if len(occupants) > 1:
                    st = slot_times.get(slot, ("?", "?"))
                    conflicts.append({
                        "room": room,
                        "day": day,
                        "slot": slot,
                        "time": f"{st[0]} - {st[1]}",
                        "entries": occupants,
                    })

    return dict(room_grids), conflicts


# ---------------------------------------------------------------------------
# CSV Output
# ---------------------------------------------------------------------------


def write_csv(room_grids, slot_times_24h, num_slots, output_path):
    """
    Write a flat room-availability CSV.
    Columns: room, day, time, occupied
    One row per (room, day, slot). Ordered by room (sorted),
    then time slot (ascending), then day (Mo-Su) within each slot.
    """
    sorted_rooms = sorted(room_grids.keys(), key=_natural_sort_key)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["room", "day", "time", "occupied"])

        for room in sorted_rooms:
            grid = room_grids[room]
            for slot in range(1, num_slots + 1):
                time_str = slot_times_24h[slot]
                for day in DAYS:
                    occupants = grid.get(day, {}).get(slot, [])
                    occupied = "true" if occupants else "false"
                    writer.writerow([room, day, time_str, occupied])

    print(f"CSV saved: {output_path}")


# ---------------------------------------------------------------------------
# Excel Output
# ---------------------------------------------------------------------------


def write_excel(room_grids, conflicts, slot_times, num_slots, output_path):
    """
    Write an Excel workbook with one sheet per room.
    Each sheet: rows = days (Mo-Su), columns = slots 1-num_slots.
    Cells contain 'Section\\nSubject'. Conflicts highlighted red.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sorted_rooms = sorted(room_grids.keys(), key=_natural_sort_key)

    # Build conflict lookup set
    conflict_set = set()
    for c in conflicts:
        conflict_set.add((c["room"], c["day"], c["slot"]))

    for room in sorted_rooms:
        grid = room_grids[room]

        # Sanitise sheet name (Excel: max 31 chars, no special chars)
        sheet_name = re.sub(r'[\\/*?\[\]:]', '_', room)[:31]
        if not sheet_name:
            sheet_name = "Unknown Room"

        # Handle duplicate sheet names
        base_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            suffix = f"_{counter}"
            sheet_name = base_name[: 31 - len(suffix)] + suffix
            counter += 1

        ws = wb.create_sheet(title=sheet_name)

        # --- Title row ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_slots + 1)
        title_cell = ws.cell(row=1, column=1, value=f"Room: {room}")
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # --- Header row (slot numbers + times) ---
        day_header = ws.cell(row=2, column=1, value="Day")
        day_header.font = HEADER_FONT
        day_header.fill = HEADER_FILL
        day_header.border = THIN_BORDER
        day_header.alignment = Alignment(horizontal="center", vertical="center")

        for slot_num in range(1, num_slots + 1):
            col = slot_num + 1
            start_t, end_t = slot_times[slot_num]
            header_text = f"{slot_num}\n{start_t}\n{end_t}"
            cell = ws.cell(row=2, column=col, value=header_text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        ws.row_dimensions[2].height = 45

        # --- Day rows ---
        for day_idx, day in enumerate(DAYS):
            row_num = day_idx + 3

            # Day label
            day_cell = ws.cell(row=row_num, column=1, value=day)
            day_cell.font = DAY_FONT
            day_cell.fill = DAY_FILL
            day_cell.border = THIN_BORDER
            day_cell.alignment = Alignment(horizontal="center", vertical="center")

            slot = 1
            while slot <= num_slots:
                col = slot + 1
                occupants = grid.get(day, {}).get(slot, [])

                if not occupants:
                    cell = ws.cell(row=row_num, column=col, value="")
                    cell.border = THIN_BORDER
                    slot += 1
                    continue

                is_conflict = (room, day, slot) in conflict_set

                # Determine span: count consecutive slots with same occupants
                span = 1
                if not is_conflict:
                    while slot + span <= num_slots:
                        next_occ = grid.get(day, {}).get(slot + span, [])
                        if next_occ == occupants:
                            span += 1
                        else:
                            break

                # Build cell text
                if is_conflict:
                    text_parts = [f"{sec}: {subj}" for sec, subj in occupants]
                    cell_text = " | ".join(text_parts)
                else:
                    sec, subj = occupants[0]
                    cell_text = f"{sec}\n{subj}"

                # Merge cells for multi-slot spans
                if span > 1:
                    ws.merge_cells(
                        start_row=row_num, start_column=col,
                        end_row=row_num, end_column=col + span - 1,
                    )

                cell = ws.cell(row=row_num, column=col, value=cell_text)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

                if is_conflict:
                    cell.fill = CONFLICT_FILL
                    cell.font = CONFLICT_FONT
                else:
                    cell.font = CELL_FONT

                # Borders on merged cells
                for s in range(1, span):
                    ws.cell(row=row_num, column=col + s).border = THIN_BORDER

                slot += span

            ws.row_dimensions[row_num].height = 50

        # --- Column widths ---
        ws.column_dimensions["A"].width = 6
        for slot_num in range(1, num_slots + 1):
            col_letter = get_column_letter(slot_num + 1)
            ws.column_dimensions[col_letter].width = 14

    # --- CONFLICTS summary sheet ---
    if conflicts:
        ws_c = wb.create_sheet(title="CONFLICTS", index=0)
        ws_c.cell(row=1, column=1, value="ROOM SCHEDULING CONFLICTS").font = Font(
            bold=True, size=14, color="FF0000"
        )
        ws_c.merge_cells("A1:F1")

        headers = ["#", "Room", "Day", "Slot", "Time", "Conflicting Classes"]
        for ci, h in enumerate(headers, 1):
            cell = ws_c.cell(row=3, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = CONFLICT_FILL
            cell.border = THIN_BORDER

        for ri, c in enumerate(conflicts, 4):
            ws_c.cell(row=ri, column=1, value=ri - 3).border = THIN_BORDER
            ws_c.cell(row=ri, column=2, value=c["room"]).border = THIN_BORDER
            ws_c.cell(row=ri, column=3, value=c["day"]).border = THIN_BORDER
            ws_c.cell(row=ri, column=4, value=c["slot"]).border = THIN_BORDER
            ws_c.cell(row=ri, column=5, value=c["time"]).border = THIN_BORDER
            classes = " | ".join(f"{sec}: {subj}" for sec, subj in c["entries"])
            ws_c.cell(row=ri, column=6, value=classes).border = THIN_BORDER

        ws_c.column_dimensions["A"].width = 5
        ws_c.column_dimensions["B"].width = 20
        ws_c.column_dimensions["C"].width = 6
        ws_c.column_dimensions["D"].width = 6
        ws_c.column_dimensions["E"].width = 15
        ws_c.column_dimensions["F"].width = 80

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        pdf_path = _get_latest_pdf(DEFAULT_INPUT_DIR)
        if pdf_path is None:
            print(f"ERROR: No PDF files found in {os.path.abspath(DEFAULT_INPUT_DIR)}")
            sys.exit(1)
    pdf_path = os.path.abspath(pdf_path)

    if not os.path.isfile(pdf_path):
        print(f"ERROR: PDF not found at {pdf_path}")
        sys.exit(1)

    # Number of pages to process (0 = all)
    max_pages = int(sys.argv[2]) if len(sys.argv) > 2 else 0

    output_path = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_OUTPUT
    output_path = os.path.abspath(output_path)

    print(f"Opening: {pdf_path}")
    print(f"Processing: {'all pages' if max_pages == 0 else f'first {max_pages} page(s)'}")
    print()

    all_entries = []

    with pdfplumber.open(pdf_path) as pdf:
        # Auto-detect time slot structure from the PDF header
        slot_times, slot_times_24h, num_slots = detect_slot_times(pdf)

        page_count = len(pdf.pages)
        pages_to_process = page_count if max_pages == 0 else min(max_pages, page_count)

        for pi in range(pages_to_process):
            page = pdf.pages[pi]
            entries = extract_page_entries(page, pi)
            all_entries.extend(entries)

            if (pi + 1) % 25 == 0 or (pi + 1) == pages_to_process:
                print(f"  Processed {pi + 1}/{pages_to_process} pages "
                      f"({len(all_entries)} entries so far)")

    print()

    if not all_entries:
        print("No entries extracted.")
        return

    rooms = set(e["room"] for e in all_entries if e["room"])
    sections = set(e["section"] for e in all_entries)
    print(f"Extracted {len(all_entries)} entries from {len(sections)} sections "
          f"across {len(rooms)} rooms")

    # Build room grids and detect conflicts
    room_grids, conflicts = build_room_grids(all_entries, slot_times, num_slots)
    print(f"Rooms with schedules: {len(room_grids)}")

    if conflicts:
        print(f"\n*** {len(conflicts)} CONFLICTS DETECTED ***")
        for c in conflicts[:10]:
            classes = " | ".join(f"{sec}: {subj}" for sec, subj in c["entries"])
            print(f"  Room {c['room']}, {c['day']} Slot {c['slot']} ({c['time']}): {classes}")
        if len(conflicts) > 10:
            print(f"  ... and {len(conflicts) - 10} more (see CONFLICTS sheet in Excel)")
    else:
        print("No conflicts detected.")

    print()

    # Write Excel
    write_excel(room_grids, conflicts, slot_times, num_slots, output_path)

    # Write CSV (room availability matrix)
    csv_output_path = os.path.abspath(DEFAULT_CSV_OUTPUT)
    write_csv(room_grids, slot_times_24h, num_slots, csv_output_path)

    print(f"\nDone! {len(all_entries)} entries -> {len(room_grids)} room sheets + CSV")


if __name__ == "__main__":
    main()
